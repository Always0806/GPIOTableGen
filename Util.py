# encoding: utf-8
import sys
import os
import signal
from openpyxl.utils import get_column_letter
from openpyxl import Workbook,load_workbook

ItemList=[]

## {{{ http://code.activestate.com/recipes/410692/ (r8)
# This class provides the functionality we want. You only need to look at
# this if you want to know how this works. It only needs to be defined
# once, no need to muck around with its internals.
class switch(object):
    def __init__(self, value):
        self.value = value
        self.fall = False
 
    def __iter__(self):
        """Return the match method once, then stop"""
        yield self.match
        raise StopIteration
     
    def match(self, *args):
        """Indicate whether or not to enter a case suite"""
        if self.fall or not args:
            return True
        elif self.value in args: # changed for v1.5, see below
            self.fall = True
            return True
        else:
            return False
			
class Items:
	def __init__(self, BallName,GPIO):
		self.BallName = BallName
		self.GPIO = GPIO
		self.NetName = None
		self.Direction = None
		self.Data = None
		self.Position = None
	
	def set_NetName(self, NetName):
		self.NetName=NetName
	
	def set_Direction(self, Direction):
		self.Direction=Direction
		
	def set_Data(self, Data):
		self.Data=Data
	
	def set_Position(self, Position):
		self.Position=Position
 
def GetCellValue(ws,index_row,column_letter):
	return ws[column_letter+str(index_row)].value
	
def GetColumnNameAndChangeValue(ws,string_row,changelist,index_row):
	string=[]
	for i in range(0,len(changelist)):
		column_letter = changelist[i]
		string.append(str(GetCellValue(ws,string_row,column_letter))+' to ['+str(GetCellValue(ws,index_row,column_letter))+'] ')
	return "".join(string)
	
def GetNumAndName(ws,index_row):
	return '['+GetCellValue(ws,index_row,'D')+'] '+GetCellValue(ws,index_row,'C')+'  : '
	
def GetColumnLetter(ws,string_row,string_value):
	for column in range(1,40):
		column_letter = get_column_letter(column)
		if ws[column_letter+str(string_row)].value==string_value:
			return column_letter
	return None

def Get_Bit(byteval,idx):
    return ((byteval&(1<<idx))!=0);
	
def AppendBit(data_L,data_M):
	output_str=""
	if data_L != 0:
		for i in range(0, 8):
			if Get_Bit(int(data_L,16),i) == True:
				output_str=output_str+str(i)+"/"
	if data_M != 0:
		for i in range(0, 8):
			if Get_Bit(int(data_M,16),i) == True:
				output_str=output_str+str(i+8)+"/"
	if data_L != 0 or data_M != 0:
		output_str=output_str+"\n"
		
	return output_str

def StringToSignint(string,len):
	x = int(string,16)
	if x > ((1<<(8*len))/2)-1:
		x -= 1<<(8*len)
	return x
	
def ExcelToStruct(filename):
	try:
		wb = load_workbook(filename)
	except IOError:
		print ("Can't open file exit")
		sys.exit(0)

	ws = wb.active
	index_row=2
	
	print ("clear All data in excel")
	tmp_row=index_row
	while True:
		BallName=ws[GetColumnLetter(ws,1,'BallName')+str(tmp_row)].value
		if BallName==None:
			break;
		
		for row in ws['C'+str(tmp_row)+':G'+str(tmp_row)]:
			for cell in row:
				cell.value = None
		tmp_row = tmp_row+1;
	
	while True:
		BallName=ws[GetColumnLetter(ws,1,'BallName')+str(index_row)].value
		if BallName==None:
			break;
			
		GPIOPPin=ws[GetColumnLetter(ws,1,'GPIO')+str(index_row)].value
		if GPIOPPin!=None:
			ItemList.append(Items(BallName,GPIOPPin))
		
		index_row = index_row+1;
		
	wb.save(filename)
		
def StructToExcel(filename):
	try:
		wb = load_workbook(filename)
	except IOError:
		print ("Can't open file exit")
		sys.exit(0)

	ws = wb.active
	index_row=2
	
	while True:
		BallName=ws[GetColumnLetter(ws,1,'BallName')+str(index_row)].value
		if BallName==None:
			break;
		
		for item in ItemList:
			if item.BallName!=None and item.NetName !=None and BallName.strip() == item.BallName.strip():
				ws[GetColumnLetter(ws,1,'NetName')+str(index_row)] = item.NetName
		
		index_row = index_row+1;
	
	wb.save(filename)

def FindBallNameAppend(BallName,Position):
	for item in ItemList:
		if BallName.strip() == item.BallName.strip():
			item.set_Position(Position)
			
def FindPositionAppend(Position,SIG_NAME):
	for item in ItemList:
		if xstr(Position).strip() == xstr(item.Position).strip():
			item.set_NetName(SIG_NAME)
def xstr(s):
    if s is None:
        return ''
    return str(s)
	
def PrintItemList():
	for item in ItemList:
		print (xstr(item.BallName)+" "+xstr(item.GPIO)+" "+xstr(item.Position)+" "+xstr(item.NetName))
